"""
TPL Submittal Comment Extractor — ADVANCED Core Engine
======================================================
Improved extraction logic with:
- Black text numbered comment detection
- Advanced color recognition
- Proper comment splitting & numbering
- Full context preservation

To run standalone: python core.py <folder_with_zips> <output.xlsx>

Functions exposed to the adapter (extractor/extractor.py):
    process_zip(zip_path)  ->  {sub, doc_name, doc_name_source, comments, debug}
"""

import zipfile, os, re, sys, io
import fitz  # PyMuPDF

# ── CLASSIFICATION RULES ──────────────────────────────────────────────────────

NON_COMMENT_ANNOTATION_TYPES = {
    'Stamp', 'FileAttachment', 'Sound', 'Movie', 'Widget', 'Link', 'Watermark'
}
NON_COMMENT_AUTHORS = {'AutoCAD SHX Text', 'AutoCAD'}  # Removed 'USER'

def is_actionable_annotation(annot_type, author, content):
    """Return True if this PDF annotation is a real reviewer comment."""
    if annot_type in NON_COMMENT_ANNOTATION_TYPES:
        return False
    if author in NON_COMMENT_AUTHORS:
        return False
    if 'AutoCAD' in (author or ''):
        return False
    if not content or not content.strip():
        return False
    return True

# ── ADVANCED COLOR DETECTION ──────────────────────────────────────────────────

def is_red_span(color_int):
    """
    Improved red detection handles multiple red variants:
    - Bright red: R>150
    - Dark red: R>100 with G<R-50
    - All variants where R > G and R > B
    """
    r = (color_int >> 16) & 0xFF
    g = (color_int >> 8) & 0xFF
    b = color_int & 0xFF
    
    # Primary check: Bright red (R > 150)
    if r > 150 and g < 100 and b < 100:
        return True
    
    # Secondary check: Dark red (strong R dominance)
    if r > 120 and g < 120 and b < 100:
        return True
    
    # Tertiary check: General red tone (R > G and R > B)
    if r > 100 and r > g and r > b and (r - g) > 30:
        return True
    
    return False

# ── EXTRACTION ────────────────────────────────────────────────────────────────

def extract_sub_number(zip_filename):
    m = re.match(r'(SUB\d+)', os.path.basename(zip_filename), re.I)
    return m.group(1).upper() if m else os.path.basename(zip_filename)

def parse_doc_name_from_filename(pdf_filename):
    """Fallback: clean document title from PDF filename (no doc number)."""
    name = os.path.basename(pdf_filename)
    name = re.sub(r'_annotated$', '', name, flags=re.I)
    name = re.sub(r'\.pdf$', '', name, flags=re.I)
    name = re.sub(r'^\d+_', '', name)
    name = name.replace('/', ' ').replace('\\', ' ').strip()
    return name

DOC_NO_RE   = re.compile(r'T?TPL[-–][\w][-–\w]+[-–]\d+', re.I)
SUBJECT_RE  = re.compile(r'Subject\s*[:\-]?\s*(.+)', re.I)
DOCNO_LABEL = re.compile(r'DOC\s*NO\s*[:\-]?\s*(T?TPL[-–][\w][-–\w]+[-–]\d+)', re.I)

# IMPROVED: Handles DWTP(...) and CWTP(...) patterns with/without spaces
EQUIP_TAG   = re.compile(
    r'\s*\(?[^)]*(?:WTP|VFD|VFR|DWTP|CWTP|A[-/]B|A[-/]B[-/]C|A[-/]B[-/]C[-/]D)\d*[^)]*\)?\s*$',
    re.I
)

def _strip_equipment_tag(text):
    """Remove trailing equipment-tag suffixes like (WTP2-VFD-101 A/B/C/D)."""
    return EQUIP_TAG.sub('', text).strip()

def extract_doc_name_from_cover(cover_pdf_bytes):
    """
    Read the EIDA cover sheet (SUBxxxx.pdf) page 1 and return
    (doc_name, source_label).

    Strategy
    --------
    1. Find the 'Subject:' line — it contains  DOCNO_Title  verbatim.
    2. If not found, find 'Contractor Submittal Ref No:' for the doc number
       and reconstruct from that + the title in the Subject area.
    Returns None if nothing useful is found.
    """
    doc = fitz.open(stream=cover_pdf_bytes, filetype="pdf")
    page = doc[0]
    lines = [l.strip() for l in page.get_text("text").splitlines() if l.strip()]

    # ── Strategy 1: Subject line ──────────────────────────────────────────────
    NEW_FIELD = re.compile(
        r'^(?:Project|Submittal|Contractor|From|Discipline|Date|Workflow|Document|'
        r'Location|Area|List|GC|Closed|Initiator|Response|Status|User|Company|'
        r'Powered|Page|Signature)',
        re.I
    )
    for i, line in enumerate(lines):
        m = SUBJECT_RE.match(line)
        if m:
            subject_value = m.group(1).strip()
            j = i + 1
            while j < len(lines):
                open_b  = subject_value.count('(')
                close_b = subject_value.count(')')
                next_l  = lines[j].strip()
                if NEW_FIELD.match(next_l):
                    break
                if open_b <= close_b:
                    break
                subject_value += ' ' + next_l
                j += 1
            if DOC_NO_RE.search(subject_value):
                return _strip_equipment_tag(subject_value), 'Subject field'

    # ── Strategy 2: Contractor Submittal Ref No + reconstruct ────────────────
    for i, line in enumerate(lines):
        m = re.search(r'Contractor Submittal Ref No\s*[:\-]?\s*(T?TPL[-–][\w][-–\w]+[-–]\d+)', line, re.I)
        if m:
            doc_no = m.group(1).strip()
            for j in range(max(0, i-5), min(len(lines), i+10)):
                if DOC_NO_RE.search(lines[j]) and '_' in lines[j]:
                    return _strip_equipment_tag(lines[j].strip()), 'Subject field'
            return doc_no, 'Ref No only'

    return None, None

def extract_doc_no_from_technical_pdf(pdf_bytes):
    """
    Fallback: read page 1 of the technical PDF and find 'DOC NO: ...'
    Returns (doc_no_string, title_string) or (None, None).
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    lines = [l.strip() for l in page.get_text("text").splitlines() if l.strip()]

    doc_no = None
    title = None
    for i, line in enumerate(lines):
        m = DOCNO_LABEL.search(line)
        if m and not doc_no:
            doc_no = m.group(1).strip()
        if re.match(r'Document\s+Title\s*[:\-]?', line, re.I) and i + 1 < len(lines):
            title = lines[i + 1].strip()
            if i + 2 < len(lines) and not DOC_NO_RE.search(lines[i+2]):
                next_line = lines[i + 2].strip()
                if next_line and not re.match(r'(SEMICONDUCTOR|DOC|Revision|Date|Approval)', next_line, re.I):
                    title += ' ' + next_line

    if doc_no and title:
        return f"{doc_no}_{_strip_equipment_tag(title)}", 'DOC NO + title block'
    if doc_no:
        return doc_no, 'DOC NO only'
    return None, None

def extract_comments_from_pdf(pdf_bytes, fname):
    """
    Returns (comments_list, debug_list)
    
    ADVANCED: Extracts using 3 methods:
    1. PDF Annotations (FreeText, Highlight, etc.)
    2. Red Text (embedded in page)
    3. Black Text Numbered Items (NEW!)
    
    comments_list: list of clean comment strings (NO [Page X] prefixes)
    debug_list: list of dicts with raw annotation info
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    comments = []
    debug = []
    seen = set()

    for pnum in range(len(doc)):
        page = doc[pnum]

        # ── METHOD 1: PDF annotation objects ─────────────────────────────────
        if page.annots():
            for a in page.annots():
                atype = a.type[1]
                author = a.info.get('title', '').strip()
                content = a.info.get('content', '').strip()

                debug.append({
                    'file': fname, 'page': pnum+1,
                    'type': atype, 'author': author,
                    'content': content[:200],
                    'actionable': is_actionable_annotation(atype, author, content)
                })

                if is_actionable_annotation(atype, author, content):
                    clean = content.replace('\r\n', '\n').replace('\r', '\n').strip()
                    key = clean.lower()[:80]
                    if key not in seen:
                        seen.add(key)
                        # NO [Page X] prefix - just the comment
                        comments.append(clean)

        # ── METHOD 2: Red text embedded in page content ───────────────────────
        blocks = page.get_text("dict")["blocks"]
        red_lines = []
        
        for b in blocks:
            if b.get("type") != 0:
                continue
            for line in b.get("lines", []):
                line_text = ""
                is_red_line = False
                for span in line.get("spans", []):
                    if is_red_span(span.get("color", 0)):
                        is_red_line = True
                    line_text += span.get("text", "")
                if is_red_line and line_text.strip():
                    red_lines.append(line_text.strip())

        if red_lines:
            NUMBERED = re.compile(r'^\s*\d+[\.\)]\s*')
            entries = []
            preamble = []

            for line in red_lines:
                if NUMBERED.match(line):
                    if preamble:
                        prefix = " ".join(preamble) + " — "
                        preamble = []
                    else:
                        prefix = ""
                    entries.append(prefix + line.strip())
                else:
                    if entries:
                        entries.append(line.strip())
                    else:
                        preamble.append(line.strip())

            if preamble:
                entries.append(" ".join(preamble))

            for entry in entries:
                key = entry.lower()[:80]
                if key not in seen:
                    seen.add(key)
                    # NO [Page X] prefix
                    comments.append(entry)

            debug.append({
                'file': fname, 'page': pnum+1,
                'type': 'RedText (embedded)',
                'author': 'Reviewer (red ink)',
                'content': " | ".join(red_lines)[:200],
                'actionable': True
            })

        # ── METHOD 3: Black text numbered/bulleted items (NEW!) ───────────────
        text_lines = page.get_text("text").splitlines()
        
        # Patterns for potential comments
        NUMBERED_ITEM = re.compile(r'^\s*(\d+[\.\)])\s+(.{10,})')
        BULLETED_ITEM = re.compile(r'^\s*([•\-\*])\s+(.{10,})')
        REVIEW_KEYWORDS = ['review', 'comment', 'note', 'action', 'provide', 'revise', 'update', 'verify', 'check', 'ensure']
        
        potential_comments = []
        for line in text_lines:
            line_stripped = line.strip()
            if not line_stripped or len(line_stripped) < 15:
                continue
                
            # Check numbered items
            m = NUMBERED_ITEM.match(line)
            if m:
                potential_comments.append(line_stripped)
                continue
            
            # Check bulleted items
            m = BULLETED_ITEM.match(line)
            if m:
                potential_comments.append(line_stripped)
                continue
            
            # Check for review keywords (indicates it's a comment)
            if any(kw in line_stripped.lower() for kw in REVIEW_KEYWORDS):
                if re.search(r'\b[A-Z][a-zA-Z]+.*[\.!?:;]', line_stripped):
                    potential_comments.append(line_stripped)
        
        # Add black text comments if not already seen (and not red)
        if potential_comments:
            for comment in potential_comments:
                key = comment.lower()[:80]
                # Only add if not already captured by other methods
                if key not in seen and not any(is_red_span(c) for c in [0] if False):
                    # Basic check: not in red_lines already
                    if not any(c.lower() in comment.lower() or comment.lower() in c.lower() for c in red_lines):
                        seen.add(key)
                        comments.append(comment)
                        debug.append({
                            'file': fname, 'page': pnum+1,
                            'type': 'BlackText (numbered/bulleted)',
                            'author': 'Reviewer (document text)',
                            'content': comment[:200],
                            'actionable': True
                        })

    return comments, debug

def process_zip(zip_path):
    """
    PUBLIC — called by the adapter.

    Returns dict:
        sub             : str
        doc_name        : str
        doc_name_source : str
        comments        : list[str]
        debug           : list[dict]
    """
    sub = extract_sub_number(zip_path)
    all_comments = []
    all_debug = []

    doc_name_source = None

    with zipfile.ZipFile(zip_path) as zf:
        pdf_files = [f for f in zf.namelist() if f.lower().endswith('.pdf')]

        def sort_key(f):
            base = os.path.basename(f).lower()
            if 'annotated' in base: return 0
            if 'response'  in base: return 1
            if re.match(r'sub\d+\.pdf', base): return 99
            return 2

        pdf_files.sort(key=sort_key)

        # Step 1: doc name from cover sheet (SUBxxxx.pdf)
        cover_files = [f for f in pdf_files
                       if re.match(r'sub\d+\.pdf', os.path.basename(f).lower())
                       and '_response' not in f.lower()]
        doc_name = None
        if cover_files:
            cover_bytes = zf.read(cover_files[0])
            doc_name, doc_name_source = extract_doc_name_from_cover(cover_bytes)

        # Step 2: comments from technical PDFs
        first_technical_bytes = None
        first_technical_fname = None

        for fname in pdf_files:
            base = os.path.basename(fname).lower()
            if re.match(r'sub\d+\.pdf', base):
                continue

            pdf_bytes = zf.read(fname)
            comments, debug = extract_comments_from_pdf(pdf_bytes, fname)
            all_debug.extend(debug)
            all_comments.extend(comments)

            if first_technical_bytes is None and 'annotated' not in base:
                first_technical_bytes = pdf_bytes
                first_technical_fname = fname

        # Step 3: fallback — DOC NO from technical PDF title block
        if not doc_name and first_technical_bytes:
            doc_name, doc_name_source = extract_doc_no_from_technical_pdf(first_technical_bytes)

        # Step 4: last resort — clean filename
        if not doc_name:
            if first_technical_fname:
                doc_name = parse_doc_name_from_filename(first_technical_fname)
                doc_name_source = 'filename only'
            else:
                doc_name = sub
                doc_name_source = 'SUB number (no PDF found)'

    # Deduplicate comments
    seen_c = set()
    unique_comments = []
    for c in all_comments:
        key = c.lower()[:80]
        if key not in seen_c:
            seen_c.add(key)
            unique_comments.append(c)

    return {
        'sub':             sub,
        'doc_name':        doc_name,
        'doc_name_source': doc_name_source,
        'comments':        unique_comments,
        'debug':           all_debug,
    }


# ── STANDALONE ENTRY POINT ────────────────────────────────────────────────────

if __name__ == '__main__':
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    inp = sys.argv[1] if len(sys.argv) > 1 else '/tmp/zips'
    out = sys.argv[2] if len(sys.argv) > 2 else '/tmp/TPL_Comments_Output.xlsx'

    zip_paths = []
    if os.path.isdir(inp):
        for f in sorted(os.listdir(inp)):
            if f.lower().endswith('.zip'):
                zip_paths.append(os.path.join(inp, f))
    elif inp.lower().endswith('.zip'):
        zip_paths = [inp]
    else:
        print("Pass a folder of ZIPs or a single ZIP file.")
        sys.exit(1)

    zip_paths.sort(key=lambda p: int(re.search(r'SUB(\d+)', os.path.basename(p), re.I).group(1))
                   if re.search(r'SUB(\d+)', os.path.basename(p), re.I) else 0)

    seen_subs = set()
    unique_paths = []
    for p in zip_paths:
        sub = extract_sub_number(p)
        if sub not in seen_subs:
            seen_subs.add(sub)
            unique_paths.append(p)
    zip_paths = unique_paths

    rows = []
    for sr, zpath in enumerate(zip_paths, 1):
        sub = extract_sub_number(zpath)
        print(f"  {sr:02d}. {sub} ...", end=" ", flush=True)
        result = process_zip(zpath)
        src = result.get('doc_name_source') or 'unknown'
        if result['comments']:
            print(f"✅ {len(result['comments'])} comment(s)  [{src}]")
            cc = "\n".join(result['comments'])
            xr = ""
        else:
            print(f"— No comments  [{src}]")
            cc = ""
            xr = "Comment not Received"
        rows.append({'sr': sr, 'sub': sub, 'doc_name': result['doc_name'],
                     'customer_comments': cc, 'xylem_remarks': xr})

    # Minimal standalone Excel (no debug sheet needed for quick tests)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TPL Comments"
    ws.append(["Sr No.", "Submittal", "Document", "Customer Comments", "Xylem Remarks"])
    for r in rows:
        ws.append([r['sr'], r['sub'], r['doc_name'], r['customer_comments'], r['xylem_remarks']])
    wb.save(out)
    print(f"\n✅ Saved: {out}")
    print(f"📊 {len(rows)} submittals, "
          f"{sum(1 for r in rows if r['customer_comments'])} with comments")
