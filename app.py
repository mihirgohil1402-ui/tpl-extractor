"""
TPL Comment Extractor — Web Application
========================================
Streamlit front-end for generating TPL Comments Excel sheets.
All extraction logic lives in extractor/extractor.py.
This file contains ONLY UI and orchestration code.
"""

import streamlit as st
import os
import sys
import time
import uuid
import shutil
import tempfile
from pathlib import Path
from datetime import datetime

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))
UPLOADS_DIR = ROOT / "uploads"
OUTPUTS_DIR = ROOT / "outputs"
UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

# ── Import extraction engine ──────────────────────────────────────────────────
try:
    from extractor import run_extraction, ExtractionResult
    ENGINE_AVAILABLE = True
except ImportError as e:
    ENGINE_AVAILABLE = False
    ENGINE_ERROR = str(e)

# ── Excel generation (pure formatting, not extraction) ────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Pandas for preview ────────────────────────────────────────────────────────
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def build_excel(rows: list, debug_rows: list, output_path: str):
    """
    Build the TPL Comments Excel file from extraction result rows.
    Matches the exact format of the hand-made TPL_Comments.xlsx.
    Also writes a Debug sheet with all raw annotation data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TPL Comments"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7    # Sr No
    ws.column_dimensions["B"].width = 12   # Submittal
    ws.column_dimensions["C"].width = 72   # Document
    ws.column_dimensions["D"].width = 62   # Customer Comments
    ws.column_dimensions["E"].width = 24   # Xylem Remarks

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Sr No.", "Submittal", "Document", "Customer Comments", "Xylem Remarks"]
    header_fill   = PatternFill("solid", fgColor="1E293B")  # dark steel
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = border

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill    = PatternFill("solid", fgColor="F8F7F4")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    cnr_fill    = PatternFill("solid", fgColor="FEF3C7")   # amber for "Comment not Received"
    cmt_fill    = PatternFill("solid", fgColor="E6F4ED")   # green for rows with comments

    base_font   = Font(name="Arial", size=9)
    sr_font     = Font(name="Arial", size=9, color="6B7280")
    sub_font    = Font(name="Arial", size=9, bold=True)
    cmt_font    = Font(name="Arial", size=9, color="0D6E3F")
    cnr_font    = Font(name="Arial", size=9, color="B45309", italic=True)

    left_align  = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    center_align= Alignment(horizontal="center", vertical="top", wrap_text=False)

    for row_idx, row in enumerate(rows, 2):
        has_comment = bool(row.get("customer_comments", "").strip())
        row_fill    = cmt_fill if has_comment else (alt_fill if row_idx % 2 == 0 else white_fill)

        # Sr No
        c = ws.cell(row=row_idx, column=1, value=row.get("sr_no", row_idx - 1))
        c.font = sr_font; c.fill = row_fill; c.alignment = center_align; c.border = border

        # Submittal
        c = ws.cell(row=row_idx, column=2, value=row.get("submittal", ""))
        c.font = sub_font; c.fill = row_fill; c.alignment = left_align; c.border = border

        # Document
        c = ws.cell(row=row_idx, column=3, value=row.get("document", ""))
        c.font = base_font; c.fill = row_fill; c.alignment = left_align; c.border = border

        # Customer Comments
        comments = row.get("customer_comments", "")
        c = ws.cell(row=row_idx, column=4, value=comments)
        c.font = cmt_font if comments else base_font
        c.fill = row_fill; c.alignment = left_align; c.border = border

        # Xylem Remarks
        remarks = row.get("xylem_remarks", "")
        c = ws.cell(row=row_idx, column=5, value=remarks)
        c.font = cnr_font if remarks == "Comment not Received" else base_font
        c.fill = cnr_fill if remarks == "Comment not Received" else row_fill
        c.alignment = left_align; c.border = border

        # Auto row height for wrapped text (approx)
        doc_lines = len(row.get("document", "").split("\n"))
        cmt_lines = max(1, len([l for l in row.get("customer_comments","").split("\n") if l.strip()]))
        ws.row_dimensions[row_idx].height = max(18, max(doc_lines, cmt_lines) * 14)

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:E{len(rows)+1}"

    # ── Debug sheet — All Annotations ────────────────────────────────────────
    if debug_rows:
        wd = wb.create_sheet("Debug - All Annotations")
        debug_headers = [
            "Submittal", "File", "Page",
            "Annotation Type", "Author", "Content (raw)", "Classified As"
        ]
        hfill  = PatternFill("solid", fgColor="1E293B")
        hfont  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        halign = Alignment(horizontal="center", vertical="center")
        for ci, h in enumerate(debug_headers, 1):
            cell = wd.cell(row=1, column=ci, value=h)
            cell.font = hfont; cell.fill = hfill
            cell.alignment = halign; cell.border = border

        for ri, row in enumerate(debug_rows, 2):
            for ci, val in enumerate(row, 1):
                cell = wd.cell(row=ri, column=ci, value=val)
                cell.font = Font(name="Arial", size=8)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

        for ci in range(1, len(debug_headers) + 1):
            wd.column_dimensions[get_column_letter(ci)].width = 16

    wb.save(output_path)


# ── CSS Styling ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Page config ── */
:root {
    --primary-dark: #1E293B;
    --primary-light: #0D6E3F;
    --accent: #FBBF24;
}

/* ── Body ── */
body { font-family: 'Inter', -apple-system, sans-serif; background: #FAFBFC; }

/* ── Header ── */
.tpl-header {
    background: linear-gradient(135deg, var(--primary-dark) 0%, #2d3748 100%);
    color: white;
    padding: 28px 24px;
    border-radius: 6px;
    margin-bottom: 24px;
}
.tpl-header .brand { font-size: 11px; font-weight: 700; letter-spacing: 0.05em; opacity: 0.7; }
.tpl-header h1 { font-size: 28px; font-weight: 700; margin: 8px 0 4px; }
.tpl-header .sub { font-size: 13px; opacity: 0.8; font-weight: 500; }

/* ── Cards ── */
.tpl-card {
    background: white;
    border-radius: 6px;
    padding: 20px;
    margin-bottom: 16px;
    border: 1px solid #E2E8F0;
}

/* ── Section labels ── */
.tpl-section-label {
    font-size: 12px;
    font-weight: 700;
    color: #475569;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 16px;
}
.tpl-section-label .dot { width: 8px; height: 8px; background: var(--primary-light); border-radius: 50%; }

/* ── File chips ── */
.file-chip {
    background: #F1F5F9;
    border-left: 3px solid var(--primary-light);
    padding: 8px 12px;
    font-size: 12px;
    margin-bottom: 6px;
    border-radius: 3px;
    font-family: 'Courier New', monospace;
}
.file-chip .ok { color: var(--primary-light); font-weight: 700; }

/* ── Stat boxes ── */
.stat-box {
    background: white;
    border-left: 3px solid #e5e7eb;
    padding: 12px;
    margin: 8px 0;
    border-radius: 3px;
    font-size: 12px;
}
.stat-box .label { color: #6b7280; font-weight: 500; }
.stat-box .num { font-size: 18px; font-weight: 700; color: #1f2937; }
.stat-box.green  .num { color: #0d6e3f; }
.stat-box.blue   .num { color: #1e40af; }
.stat-box.orange .num { color: #d97706; }
.stat-box.amber .num { color: #b45309; }
.stat-box.red   .num { color: #c0392b; }

/* ── Info note ── */
.info-note {
    background: #e8f0fe;
    border-left: 3px solid #1a56db;
    border-radius: 0 4px 4px 0;
    padding: 10px 14px;
    font-size: 12px;
    color: #1e40af;
    margin-bottom: 16px;
    line-height: 1.6;
}

/* ── Error note ── */
.error-note {
    background: #fdecea;
    border-left: 3px solid #c0392b;
    border-radius: 0 4px 4px 0;
    padding: 10px 14px;
    font-size: 12px;
    color: #c0392b;
    margin-bottom: 10px;
    line-height: 1.6;
}

/* ── Buttons ── */
.stButton > button {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    border-radius: 4px !important;
    transition: opacity 0.15s !important;
}
.stButton > button:hover { opacity: 0.88 !important; }

/* ── Download button ── */
.stDownloadButton > button {
    background: #0d6e3f !important;
    color: white !important;
    font-weight: 700 !important;
    border: none !important;
    padding: 12px 24px !important;
    font-size: 14px !important;
    border-radius: 4px !important;
    width: 100%;
}

/* ── Log box ── */
.log-box {
    background: #0d1117;
    border-radius: 4px;
    padding: 12px 14px;
    font-family: 'Courier New', monospace;
    font-size: 11px;
    color: #94a3b8;
    max-height: 180px;
    overflow-y: auto;
    line-height: 1.9;
    margin-top: 12px;
}

/* ── Hide streamlit branding ── */
#MainMenu, footer, [data-testid="stToolbar"] { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="tpl-header">
  <div>
    <div class="brand">TPL · Dholera Fab 1</div>
    <h1>Generate TPL Comment Sheets Automatically</h1>
    <div class="sub">Upload ZIP packages → Extract Comments → Download Excel</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ── Engine check ──────────────────────────────────────────────────────────────
if not ENGINE_AVAILABLE:
    st.markdown(f"""
    <div class="error-note">
        <strong>⚠ Extraction engine not found.</strong><br/>
        Make sure <code>extractor/extractor.py</code> exists in the project folder.<br/>
        Error: {ENGINE_ERROR}
    </div>
    """, unsafe_allow_html=True)
    st.stop()

if not EXCEL_AVAILABLE:
    st.markdown("""
    <div class="error-note">
        <strong>⚠ openpyxl not installed.</strong><br/>
        Run: <code>pip install openpyxl</code>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ── Session state ─────────────────────────────────────────────────────────────
if "result"       not in st.session_state: st.session_state.result       = None
if "excel_bytes"  not in st.session_state: st.session_state.excel_bytes  = None
if "excel_name"   not in st.session_state: st.session_state.excel_name   = None
if "elapsed"      not in st.session_state: st.session_state.elapsed      = None
if "log_lines"    not in st.session_state: st.session_state.log_lines    = []


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="tpl-card">', unsafe_allow_html=True)
st.markdown('<div class="tpl-section-label"><span class="dot"></span>Step 1 — Upload ZIP Files</div>', unsafe_allow_html=True)

st.markdown("""
<div class="info-note">
  Upload the ZIP packages received from the submittal workflow.<br/>
  Each ZIP file = one submittal row in the output Excel.
  You can select multiple ZIPs at once (hold Ctrl or Cmd).
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    label="Drop ZIP files here or click to browse",
    type=["zip"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key="file_uploader",
)
st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — UPLOADED FILES LIST
# ═══════════════════════════════════════════════════════════════════════════════
if uploaded_files:
    st.markdown('<div class="tpl-card">', unsafe_allow_html=True)
    st.markdown('<div class="tpl-section-label"><span class="dot"></span>Step 2 — Uploaded Files</div>', unsafe_allow_html=True)

    chips_html = ""
    for f in uploaded_files:
        size_kb = f.size // 1024
        chips_html += f'<div class="file-chip"><span class="ok">▪</span> {f.name} <span style="color:#9ca3af;font-size:10px;">({size_kb} KB)</span></div>'

    st.markdown(chips_html, unsafe_allow_html=True)
    st.markdown(f"<br/><strong style='font-size:13px;'>Total: {len(uploaded_files)} ZIP file{'s' if len(uploaded_files)!=1 else ''} ready to process</strong>", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — PROCESS
# ═══════════════════════════════════════════════════════════════════════════════
if uploaded_files:
    st.markdown('<div class="tpl-card">', unsafe_allow_html=True)
    st.markdown('<div class="tpl-section-label"><span class="dot"></span>Step 3 — Generate Excel</div>', unsafe_allow_html=True)

    if st.button("⚡ Process & Generate Excel", use_container_width=True):
        progress_placeholder = st.empty()
        log_placeholder = st.empty()

        with progress_placeholder.container():
            progress_bar = st.progress(0, text="Starting extraction...")

        start_time = time.time()
        log_entries = []

        def log_msg(msg: str):
            log_entries.append(msg)
            with log_placeholder.container():
                st.markdown(f'<div class="log-box">{"<br/>".join(log_entries[-20:])}</div>', unsafe_allow_html=True)

        try:
            log_msg("📋 Saving uploaded ZIPs to disk...")
            temp_uploads = {}
            for f in uploaded_files:
                temp_path = UPLOADS_DIR / f.name
                with open(temp_path, "wb") as fh:
                    fh.write(f.getbuffer())
                temp_uploads[f.name] = temp_path
                log_msg(f"  ✓ {f.name}")

            log_msg("\n🔍 Running extraction engine...")
            progress_bar.progress(33, text="Extracting comments...")

            result = run_extraction(list(temp_uploads.values()), debug=True)
            st.session_state.result = result

            log_msg(f"  ✓ Processed {len(temp_uploads)} file(s)")
            log_msg(f"  ✓ Found {len(result.rows)} rows")

            log_msg("\n📝 Generating Excel file...")
            progress_bar.progress(66, text="Building Excel...")

            output_filename = f"TPL_Comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = OUTPUTS_DIR / output_filename

            build_excel(result.rows, result.debug, str(output_path))

            with open(output_path, "rb") as fh:
                st.session_state.excel_bytes = fh.read()
                st.session_state.excel_name = output_filename

            elapsed = time.time() - start_time
            st.session_state.elapsed = elapsed

            log_msg(f"  ✓ Excel saved: {output_filename}")
            log_msg(f"\n✅ Complete! ({elapsed:.1f}s)")

            progress_bar.progress(100, text="Done!")

            st.success(f"✅ Extraction complete in {elapsed:.1f}s")

        except Exception as e:
            log_msg(f"\n❌ ERROR: {str(e)}")
            st.error(f"Extraction failed: {str(e)}")

    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — PREVIEW
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.result:
    st.markdown('<div class="tpl-card">', unsafe_allow_html=True)
    st.markdown('<div class="tpl-section-label"><span class="dot"></span>Step 4 — Preview</div>', unsafe_allow_html=True)

    if PANDAS_AVAILABLE:
        preview_data = []
        for row in st.session_state.result.rows:
            doc = row.get("document", "") or ""
            cmt = row.get("customer_comments","") or ""
            doc_short = doc[:80] + "…" if doc and len(doc) > 80 else (doc or "")
            cmt_short = cmt[:60] + "…" if cmt and len(cmt) > 60 else (cmt or "")
            preview_data.append({
                "Sr No": row.get("sr_no",""),
                "Submittal": row.get("submittal",""),
                "Document": doc_short.replace("\n"," | "),
                "Comments": cmt_short or "—",
                "Xylem Remarks": row.get("xylem_remarks",""),
            })

        df = pd.DataFrame(preview_data)
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Sr No":        st.column_config.NumberColumn(width="small"),
                "Submittal":    st.column_config.TextColumn(width="small"),
                "Document":     st.column_config.TextColumn(width="large"),
                "Comments":     st.column_config.TextColumn(width="large"),
                "Xylem Remarks":st.column_config.TextColumn(width="medium"),
            }
        )

    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — DOWNLOAD
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.excel_bytes:
    st.markdown('<div class="tpl-card">', unsafe_allow_html=True)
    st.markdown('<div class="tpl-section-label"><span class="dot"></span>Step 5 — Download</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="info-note">
      Your Excel file is ready. Click below to download.<br/>
      <strong>Review the file before sharing</strong> — the extracted comments are a draft for human review.
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label=f"⬇  Download {st.session_state.excel_name}",
        data=st.session_state.excel_bytes,
        file_name=st.session_state.excel_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — STORAGE MANAGEMENT (Footer)
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("### 💾 Storage Management")

# Calculate storage
def get_storage_size():
    total = 0
    if UPLOADS_DIR.exists():
        for f in UPLOADS_DIR.rglob("*"):
            if f.is_file(): total += f.stat().st_size
    if OUTPUTS_DIR.exists():
        for f in OUTPUTS_DIR.rglob("*"):
            if f.is_file(): total += f.stat().st_size
    return total

storage_mb = get_storage_size() / (1024 * 1024)
if storage_mb >= 1024:
    st.write(f"📊 **Current storage:** {storage_mb / 1024:.2f} GB")
else:
    st.write(f"📊 **Current storage:** {storage_mb:.2f} MB")

if st.button("🗑️ Clear All Storage", use_container_width=True):
    try:
        if UPLOADS_DIR.exists():
            shutil.rmtree(UPLOADS_DIR)
            UPLOADS_DIR.mkdir(exist_ok=True)
        if OUTPUTS_DIR.exists():
            shutil.rmtree(OUTPUTS_DIR)
            OUTPUTS_DIR.mkdir(exist_ok=True)
        st.session_state.result = None
        st.session_state.excel_bytes = None
        st.session_state.excel_name = None
        st.success("✅ Storage cleared! Ready for new uploads.")
        st.rerun()
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="text-align:center;padding:32px 0 8px;font-size:11px;color:#9ca3af;">
  TPL Comment Extractor · Internal Tool · Dholera Fab 1 Project<br/>
  Generated Excel is a draft for human review — verify before submission.
</div>
""", unsafe_allow_html=True)
